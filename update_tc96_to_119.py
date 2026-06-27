import openpyxl
import re

file_path = 'test_plan.xlsx'
wb = openpyxl.load_workbook(file_path)
sheet = wb.active

login_updates = {
    r'Login as Head': r"Login as user 'proc_head1' with password 'proc_head1' (Role: Procurement Head 1)",
    r'Login as Staff': r"Login as user 'procurement1' with password 'procurement1' (Role: Procurement1)",
    r'Login as Admin': r"Login as user 'super_admin' with password 'super_admin' (Role: Super Admin)"
}

text_updates = {
    'Projects exist in WAITING_CANCEL status': 'Project: "User Testing - Air purifier order"',
    'Assigned tab contains projects with various statuses': 'Observe projects like "User Testing - Completed tablet procurement" and "User Testing - Network monitoring sensors"',
    'Project in WAITING_ACCEPT status': 'Project: "User Testing - Replacement laptops" (WAITING_ACCEPT)',
    'Multiple projects in WAITING_ACCEPT status': 'Projects like "User Testing - Replacement laptops" and "User Testing - Desktop upgrades"',
    'Unassigned project': 'Project: "User Testing - New chairs for reading room" (UNASSIGNED)',
    'Multiple unassigned projects': 'Projects: "User Testing - New chairs for reading room" and "User Testing - Whiteboard markers"',
    'Projects in WAITING_ACCEPT and IN_PROGRESS statuses': 'Projects like "User Testing - Replacement laptops" and "User Testing - Network monitoring sensors"',
    'Valid cancel reason': 'Project: "User Testing - Network monitoring sensors" with reason "Duplicate plan"',
    'WAITING_ACCEPT': 'รอรับทราบ (WAITING_ACCEPT)',
    'IN_PROGRESS': 'มอบหมายแล้ว (IN_PROGRESS)',
    'WAITING_CANCEL': 'รออนุมัติยกเลิก (WAITING_CANCEL)',
    'UNASSIGNED': 'ยังไม่ได้มอบหมาย (UNASSIGNED)',
    'CANCELLED': 'ยกเลิก (CANCELLED)'
}

# The expected values are a bit delicate. Let's do string replacement carefully.
# In openpyxl, rows and cols are 1-indexed.
# Usually:
# Col A (1) = TC ID
# Col E (5) = Test Data
# Col F (6) = Steps
# Col G (7) = Expected Result

modified_count = 0

for row_idx in range(2, sheet.max_row + 1):
    tc_id_cell = sheet.cell(row=row_idx, column=1)
    if not tc_id_cell.value or not str(tc_id_cell.value).startswith('TC-'):
        continue
    
    try:
        tc_num = int(str(tc_id_cell.value).replace('TC-', ''))
    except ValueError:
        continue
        
    if 96 <= tc_num <= 119:
        row_modified = False
        
        # Update Steps (Col 6)
        step_cell = sheet.cell(row=row_idx, column=6)
        if step_cell.value:
            original = str(step_cell.value)
            new_val = original
            for pattern, repl in login_updates.items():
                new_val = re.sub(pattern, repl, new_val, flags=re.IGNORECASE)
            if new_val != original:
                step_cell.value = new_val
                row_modified = True

        # Update Data (Col 5) and Expected (Col 7)
        for col_idx in [5, 7]:
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value:
                original = str(cell.value)
                new_val = original
                for old_str, new_str in text_updates.items():
                    if old_str in new_val:
                        # Prevent double replacement if script is run twice
                        if new_str not in new_val:
                            new_val = new_val.replace(old_str, new_str)
                if new_val != original:
                    cell.value = new_val
                    row_modified = True
                    
        if row_modified:
            modified_count += 1

wb.save(file_path)
print(f"Updated {modified_count} rows in {file_path} using openpyxl.")
